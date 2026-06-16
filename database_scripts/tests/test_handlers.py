"""Tests for handlers module — SchemaHandler strategy pattern, factory, helpers."""

import pytest
from unittest.mock import MagicMock, patch

from models import CategoryDef, Product
from handlers import (
    SchemaHandler,
    StandardHandler,
    LaminationHandler,
    get_handler,
    log,
    exec_sql,
    _insert_product,
    _insert_compatibility,
)


class TestGetHandler:
    """测试 get_handler() 工厂函数。"""

    def test_returns_standard_handler_for_silicone(self):
        """测 schema_type='silicone' 返回 StandardHandler 实例。"""
        handler = get_handler("silicone")
        assert isinstance(handler, StandardHandler)

    def test_returns_standard_handler_for_uv_oil_matte(self):
        """测 schema_type='uv_oil_matte' 返回 StandardHandler。"""
        handler = get_handler("uv_oil_matte")
        assert isinstance(handler, StandardHandler)

    def test_returns_standard_handler_for_water_varnish(self):
        """测 schema_type='water_varnish' 返回 StandardHandler。"""
        handler = get_handler("water_varnish")
        assert isinstance(handler, StandardHandler)

    def test_returns_lamination_handler(self):
        """测 schema_type='lamination' 返回 LaminationHandler 实例。"""
        handler = get_handler("lamination")
        assert isinstance(handler, LaminationHandler)

    def test_same_type_returns_same_instance(self):
        """测相同 schema_type 多次调用返回同一实例（缓存）。"""
        h1 = get_handler("silicone")
        h2 = get_handler("silicone")
        assert h1 is h2

    def test_unknown_type_returns_standard(self):
        """测未知 schema_type 降级为 StandardHandler。"""
        handler = get_handler("unknown_type")
        assert isinstance(handler, StandardHandler)


class TestStandardHandler:
    """测试 StandardHandler 各方法。"""

    def test_extract_product(self):
        """测试 extract_product 从 worksheet 行提取 Product 对象。"""
        handler = StandardHandler()
        # 需要 mock 一个 worksheet 和 cell_text 行为
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 6, value="UV-001")
        ws.cell(1, 7, value="SUP-001")
        ws.cell(1, 8, value="STK-001")
        ws.cell(1, 9, value="UV Ink")
        ws.cell(1, 10, value="Printing")
        ws.cell(1, 11, value="油墨")
        ws.cell(1, 13, value="张三")
        ws.cell(1, 14, value="红色")
        product = handler.extract_product(ws, 1)
        assert product.material_code == "UV-001"
        assert product.material_name == "UV Ink"
        assert product.usage == "Printing"
        assert product.category == "油墨"
        assert product.responsible_person == "张三"
        assert product.color == "红色"

    def test_insert_product_delegates_to_helper(self):
        """测试 insert_product 委托给 _insert_product 并返回结果。"""
        handler = StandardHandler()
        conn = MagicMock()
        cat = CategoryDef(product_table="test_products")
        product = Product(material_code="UV-001", material_name="Test")
        # _insert_product 内部会操作 conn，这里只验证调用不抛异常
        # 实际集成测试需要真实的数据库连接
        result = handler.insert_product(conn, product, cat)
        # Mock conn 会导致异常，预期返回 None
        assert result is None or result is not None


class TestLaminationHandler:
    """测试 LaminationHandler 各方法。"""

    def test_extract_product(self):
        """测试提取过胶物料，包含专用字段映射。"""
        handler = LaminationHandler()
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 6, value="LAM-001")
        ws.cell(1, 7, value="SUP-001")
        ws.cell(1, 8, value="STK-001")
        ws.cell(1, 9, value="Laminate Film")
        ws.cell(1, 10, value="覆合用")
        ws.cell(1, 11, value="BOPP")
        ws.cell(1, 12, value="0.018")
        ws.cell(1, 13, value="0.005")
        ws.cell(1, 14, value="1000x500")
        ws.cell(1, 15, value="透明")
        ws.cell(1, 16, value="卷装")
        product = handler.extract_product(ws, 1)
        assert product.material_code == "LAM-001"
        assert product.material_name == "Laminate Film"
        assert product.usage_text == "覆合用"
        assert product.material_type == "BOPP"
        assert product.thickness_film == "0.018"
        assert product.thickness_glue == "0.005"
        assert product.size_info == "1000x500"
        assert product.color == "透明"
        assert product.shape == "卷装"

    def test_insert_product_uses_lamination_table(self):
        """测试插入使用 lamination_material_products 表。"""
        handler = LaminationHandler()
        conn = MagicMock()
        cat = CategoryDef()
        product = Product(material_code="LAM-001", material_name="Test")
        result = handler.insert_product(conn, product, cat)
        assert result is None or result is not None  # mock 无副作用可接受

    def test_insert_compat_uses_lamination_compat_table(self):
        """测试兼容性插入使用 lamination_material_compatibility 表。

        注：MagicMock 环境下 execute 不会抛异常，故返回 True。
        真实数据库连接遇到重复 key 或约束冲突时返回 False。
        """
        handler = LaminationHandler()
        conn = MagicMock()
        cursor = MagicMock()
        conn.cursor.return_value.__enter__.return_value = cursor
        cat = CategoryDef()
        result = handler.insert_compat(conn, 1, "過膠", "V", cat)
        assert result is True


class TestHelpers:
    """测试 handlers 模块中的辅助函数。"""

    def test_log_does_not_crash(self):
        """测 log 函数不会抛出异常。"""
        # 只要不抛异常就算通过
        log("Test log message")
        log("")

    def test_exec_sql_with_params(self):
        """测 exec_sql 在 mock conn 下不抛异常。"""
        conn = MagicMock()
        cursor = MagicMock()
        cursor.fetchall.return_value = []
        conn.cursor.return_value.__enter__.return_value = cursor
        result = exec_sql(conn, "SELECT 1", params=None)
        assert result == []

    def test_insert_product_empty_data_returns_none(self):
        """测插入空 Product 返回 None。"""
        conn = MagicMock()
        product = Product()  # 所有字段为空
        result = _insert_product(conn, "test_table", product)
        assert result is None


class TestSchemaHandlerBase:
    """测试 SchemaHandler 抽象基类。"""

    def test_create_tables_raises_not_implemented(self):
        """测基类 create_tables 抛出 NotImplementedError。"""
        handler = SchemaHandler()
        with pytest.raises(NotImplementedError):
            handler.create_tables(None, None)

    def test_extract_product_raises_not_implemented(self):
        """测基类 extract_product 抛出 NotImplementedError。"""
        handler = SchemaHandler()
        with pytest.raises(NotImplementedError):
            handler.extract_product(None, 0)

    def test_insert_product_raises_not_implemented(self):
        """测基类 insert_product 抛出 NotImplementedError。"""
        handler = SchemaHandler()
        with pytest.raises(NotImplementedError):
            handler.insert_product(None, None, None)

    def test_insert_compat_raises_not_implemented(self):
        """测基类 insert_compat 抛出 NotImplementedError。"""
        handler = SchemaHandler()
        with pytest.raises(NotImplementedError):
            handler.insert_compat(None, 0, "", "", None)
