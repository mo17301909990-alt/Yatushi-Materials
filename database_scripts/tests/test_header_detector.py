"""Tests for header_detector module — cell helpers, header location, data extraction."""

import pytest
import openpyxl
from header_detector import HeaderDetector, HeaderConfig


# ── Low-level cell helpers ────────────────────────────────────────────────


class TestCellText:
    """测试 cell_text() 安全获取单元格文本。"""

    def test_empty_cell_returns_empty_string(self, detector):
        """测空单元格返回空字符串。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert detector.cell_text(ws, 1, 1) == ""

    def test_none_cell_returns_empty_string(self, detector):
        """测单元格值为 None 时返回空字符串。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, value=None)
        assert detector.cell_text(ws, 1, 1) == ""

    def test_string_cell_returns_trimmed_text(self, detector):
        """测有文本的单元格返回去除前后空白的文本。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, value="  單面印刷  ")
        assert detector.cell_text(ws, 1, 1) == "單面印刷"

    def test_numeric_cell_converts_to_string(self, detector):
        """测数值类型单元格转为字符串返回。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, value=123)
        assert detector.cell_text(ws, 1, 1) == "123"


class TestNormalizeMultiline:
    """测试 normalize_multiline() 多行文本格式统一。"""

    def test_normalize_newlines_to_slash(self, detector):
        """测 \\n 换行符替换为 ' / '。"""
        assert detector.normalize_multiline("a\nb") == "a / b"

    def test_normalize_literal_backslash_n(self, detector):
        """测 \\\\n 字面量替换为 ' / '。"""
        assert detector.normalize_multiline("a\\nb") == "a / b"

    def test_triple_newline_compressed(self, detector):
        """测连续多个换行被压缩为一个 ' / '。"""
        result = detector.normalize_multiline("a\n\n\nb")
        # "a /  /  / b" -> 压缩后 "a / b"
        assert " / " in result
        assert result.count(" / ") == 1

    def test_none_returns_empty(self, detector):
        """测传入 None 返回空字符串。"""
        assert detector.normalize_multiline(None) == ""

    def test_empty_string_returns_empty(self, detector):
        """测空字符串返回空字符串。"""
        assert detector.normalize_multiline("") == ""


class TestIsValidStatus:
    """测试 is_valid_status() 兼容状态校验。"""

    def test_valid_V(self, detector):
        """测 'V' 为有效状态。"""
        assert detector.is_valid_status("V") is True

    def test_valid_X(self, detector):
        """测 'X' 为有效状态。"""
        assert detector.is_valid_status("X") is True

    def test_valid_triangle(self, detector):
        """测 '▷' 为有效状态。"""
        assert detector.is_valid_status("▷") is True

    def test_valid_O(self, detector):
        """测 'O' 为有效状态。"""
        assert detector.is_valid_status("O") is True

    def test_invalid_empty(self, detector):
        """测空字符串为无效状态。"""
        assert detector.is_valid_status("") is False

    def test_invalid_two_chars(self, detector):
        """测 'VX' 双字符为无效状态。"""
        assert detector.is_valid_status("VX") is False

    def test_invalid_lowercase(self, detector):
        """测小写 'v' 为无效状态（需严格大写）。"""
        assert detector.is_valid_status("v") is False

    def test_invalid_random_text(self, detector):
        """测随意文本为无效状态。"""
        assert detector.is_valid_status("maybe") is False

    def test_invalid_whitespace_V(self, detector):
        """测 ' V ' 前后空白时仍为有效（strip 后）"""
        assert detector.is_valid_status(" V ") is True

    def test_invalid_none(self, detector):
        """测 None 为无效状态。"""
        assert detector.is_valid_status(None) is False


# ── Sheet / header locating ──────────────────────────────────────────────


class TestFindSheet:
    """测试 find_sheet() 工作表定位。"""

    def test_prefers_applicability_sheet(self, detector, sample_workbook):
        """测优先匹配名称含'適用性'的工作表。"""
        ws = sample_workbook["適用性測試"]
        result = detector.find_sheet(ws)
        assert result.title == "適用性測試"

    def test_falls_back_to_largest_sheet(self, detector, edge_case_workbook):
        """测无适用性 sheet 时选择行列数足够的工作表。"""
        ws = edge_case_workbook.active
        result = detector.find_sheet(ws)
        # max_row=1, max_column=1 < threshold, 返回原 ws
        assert result is ws


class TestFindStepHeaderRow:
    """测试 find_step_header_row() 步骤表头行定位。"""

    def test_finds_step_row(self, detector, sample_workbook):
        """测能在 sample_workbook 的 row 18 中找到步骤表头行（默认 row_start=15）。"""
        ws = sample_workbook["適用性測試"]
        row = detector.find_step_header_row(ws)
        assert row == 18

    def test_no_step_row_returns_none(self, detector, edge_case_workbook):
        """测工作表中没有步骤关键词时返回 None。"""
        ws = edge_case_workbook.active
        row = detector.find_step_header_row(ws)
        assert row is None

    def test_with_custom_config(self, custom_config_detector, sample_workbook):
        """测使用自定义 HeaderConfig 也能正确定位步骤行。"""
        ws = sample_workbook["適用性測試"]
        row = custom_config_detector.find_step_header_row(ws)
        assert row is not None


class TestFindCategoryHeaderRow:
    """测试 find_category_header_row() 分类组表头行定位。"""

    def test_finds_category_row(self, detector, sample_workbook):
        """测能在 sample_workbook 的 row 14 中找到分类组表头行。"""
        ws = sample_workbook["適用性測試"]
        row = detector.find_category_header_row(ws)
        assert row == 14

    def test_no_category_row_returns_fallback(self, detector):
        """测步骤行上方没有分类关键词时返回 step_row - 1。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(15, 40, value="單面印刷")
        ws.cell(15, 41, value="雙面印刷")
        row = detector.find_category_header_row(ws)
        assert row == 14  # step_row - 1


# ── Data row / pair extraction ───────────────────────────────────────────


class TestFindDataRows:
    """测试 find_data_rows() 数据行提取。"""

    def test_finds_data_rows(self, detector, sample_workbook):
        """测能正确提取 sample_workbook 中的有效数据行（跳过空行）。"""
        ws = sample_workbook["適用性測試"]
        rows = detector.find_data_rows(ws, step_row=18)
        assert len(rows) == 2
        # 第一行是主物料
        assert rows[0].is_primary is True
        assert rows[0].code == "#UV-001"
        assert rows[0].name == "UV油墨測試材料"
        assert rows[0].row == 19
        # 第二行是普通物料
        assert rows[1].is_primary is False
        assert rows[1].code == "YZ-123"
        assert rows[1].name == "啞光油"
        assert rows[1].row == 20

    def test_skips_empty_cells(self, detector, sample_workbook):
        """测 col 6 或 col 9 为空的数据行被跳过（row 21）。"""
        ws = sample_workbook["適用性測試"]
        rows = detector.find_data_rows(ws, step_row=18)
        codes = [r.code for r in rows]
        assert "" not in codes

    def test_empty_worksheet_returns_empty(self, detector):
        """测空工作表返回空列表。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = detector.find_data_rows(ws, step_row=5)
        assert rows == []


class TestExtractCompatibilityPairs:
    """测试 extract_compatibility_pairs() 兼容性数据对提取。"""

    def test_extracts_pairs(self, detector, sample_workbook):
        """测从物料行正确提取 (step_name, status) 对。"""
        ws = sample_workbook["適用性測試"]
        pairs = detector.extract_compatibility_pairs(ws, row_num=19, step_row=18, cat_row=14)
        assert len(pairs) == 9  # 9 个步骤列，每列一个值
        # 验证第一个 pair 包含分类前缀
        first_step, first_status = pairs[0]
        assert "印刷" in first_step
        assert first_status in ("V", "X")
        # 验证只包含 V/X/▷/O
        for _, status in pairs:
            assert status in ("V", "X", "▷", "O"), f"Unexpected status: {status}"

    def test_no_cat_row_uses_step_name_only(self, detector, sample_workbook):
        """测没有分类行时步骤名不加前缀。"""
        ws = sample_workbook["適用性測試"]
        pairs = detector.extract_compatibility_pairs(ws, row_num=19, step_row=18, cat_row=None)
        assert len(pairs) == 9
        step_name, _ = pairs[0]
        assert step_name == "單面印刷"  # 无前缀

    def test_empty_cells_skipped(self, detector):
        """测单元格值为空时跳过不生成 pair。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(9, 40, value="")    # 空值
        ws.cell(9, 41, value="V")   # 有效值，但步骤行该列为空
        ws.cell(8, 41, value="")    # 步骤名称为空
        pairs = detector.extract_compatibility_pairs(ws, row_num=9, step_row=8)
        assert pairs == []


# ── Initialization ────────────────────────────────────────────────────────


class TestDetectorInit:
    """测试 HeaderDetector 初始化行为。"""

    def test_default_config(self):
        """测无参数时使用 HeaderConfig 默认值。"""
        d = HeaderDetector()
        assert d.config.row_start == 15
        assert d.config.row_end == 30
        assert "單面印刷" in d.config.step_keywords
        assert "印刷" in d.config.category_keywords

    def test_custom_config(self):
        """测传入自定义配置覆盖默认值。"""
        cfg = HeaderConfig(row_start=1, row_end=5)
        d = HeaderDetector(config=cfg)
        assert d.config.row_start == 1
        assert d.config.row_end == 5
