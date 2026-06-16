"""pytest fixtures for database_scripts tests."""

import pytest
import openpyxl
from header_detector import HeaderDetector, HeaderConfig


@pytest.fixture
def detector():
    """提供默认配置的 HeaderDetector 实例。"""
    return HeaderDetector()


@pytest.fixture
def custom_config_detector():
    """提供自定义测试配置的 HeaderDetector 实例，检测范围更紧凑。"""
    config = HeaderConfig(
        step_keywords={"單面印刷", "雙面印刷", "燙金"},
        category_keywords={"印刷", "燙金", "物料"},
        row_start=14,
        row_end=25,
        col_start=38,
        col_end=50,
    )
    return HeaderDetector(config)


@pytest.fixture
def sample_workbook():
    """
    创建一个最小化的 xlsx 用于测试表头检测。

    布局与真实文件对齐：分类组表头在第 14 行，步骤表头在第 18 行，
    列位置在 40+，匹配默认 HeaderConfig 的 row_start=15/col_start=40 扫描范围。

    布局：
      - Row 1-13: 空白
      - Row 14:   分类组表头 -- 列 40-42: 印刷, 43-45: 燙金, 46-48: 過膠
      - Row 15-17: 空白（非表头行）
      - Row 18:   步骤表头 -- 9 个工序名称
      - Row 19:   首行物料数据（主物料 #UV-001）
      - Row 20:   第二行物料数据
      - Row 21:   无效行（col 6 为空）

    Sheet 名称为"適用性測試"，可被 find_sheet() 优先匹配。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "適用性測試"

    CAT_ROW = 14
    STEP_ROW = 18
    DATA_ROW_1 = 19
    DATA_ROW_2 = 20
    EMPTY_ROW = 21

    STEP_HEADERS = [
        "單面印刷", "雙面印刷", "連線印刷",
        "普通燙金", "鐳射燙金", "立體燙金",
        "過膠", "絲印", "植毛",
    ]

    # ── Row 14: 分类组表头 ──
    for c in range(40, 43):
        ws.cell(CAT_ROW, c, value="印刷")
    for c in range(43, 46):
        ws.cell(CAT_ROW, c, value="燙金")
    for c in range(46, 49):
        ws.cell(CAT_ROW, c, value="過膠")

    # ── Row 18: 步骤表头 ──
    for i, h in enumerate(STEP_HEADERS):
        ws.cell(STEP_ROW, 40 + i, value=h)

    # ── Row 19: 数据行（主物料 #UV-001） ──
    ws.cell(DATA_ROW_1, 6, value="#UV-001")   # 物料编号
    ws.cell(DATA_ROW_1, 9, value="UV油墨測試材料")  # 物料名称
    for c in range(40, 49):
        ws.cell(DATA_ROW_1, c, value="V" if c in (40, 41, 43, 46) else "X")

    # ── Row 20: 数据行（从物料 YZ-123） ──
    ws.cell(DATA_ROW_2, 6, value="YZ-123")
    ws.cell(DATA_ROW_2, 9, value="啞光油")
    for c in range(40, 49):
        ws.cell(DATA_ROW_2, c, value="X")

    # ── Row 21: 无效行（col 6 为空） ──
    ws.cell(EMPTY_ROW, 6, value="")
    ws.cell(EMPTY_ROW, 9, value="跳過此行")

    return wb


@pytest.fixture
def edge_case_workbook():
    """
    边界情况 workbook：没有适用性 sheet、行列不足。

    用于测试 find_sheet() 在无匹配时的降级行为。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, value="Minimal data only")
    return wb
