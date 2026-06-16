#!/usr/bin/env python3
"""
Export ALL P0 module data from PostgreSQL to individual Excel files.
One workbook per module, with "产品列表" and "兼容性矩阵" sheets.

Usage:
    python database_scripts/export_to_excel.py

Output:
    F:/YTS+JYY/审核导出/{category_key}_审核.xlsx  (15 files)
    F:/YTS+JYY/审核导出/00_汇总清单.xlsx
    F:/YTS+JYY/审核导出/00_审核导出报告.md
"""

import os
import sys
from collections import OrderedDict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# DB connection
# ---------------------------------------------------------------------------
DB_CONFIG = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "gold_foil_db"),
    user=os.environ.get("PGUSER", "postgres"),
    password=os.environ.get("PGPASSWORD", "HryENprJrxThYSDz"),
)

# ---------------------------------------------------------------------------
# Module definition
# ---------------------------------------------------------------------------
# For each module:
#   key         -> used in file naming and table name derivation
#   display     -> Chinese label for the summary sheet
#   product_tbl -> product table name
#   compat_tbl  -> compatibility table name
#   usage_col   -> the "usage" / "usage_text" column name on the product table
#
# Rules:
#   - lamination_material uses "usage_text"; all others use "usage"
#   - Tables named: {key}_product / {key}_compatibility
#     EXCEPT lamination which is lamination_material_products / lamination_material_compatibility

MODULES = [
    # --- Silicone (7 with actual data) ---
    dict(key="silicone_glow_ink",                display="硅胶发光油墨"),
    dict(key="silicone_white_uv",                 display="硅胶白UV"),
    dict(key="silicone_glittering_star",          display="硅胶闪星"),
    dict(key="silicone_screen_uv",                display="硅胶丝印UV"),
    dict(key="silicone_led_uv_spray",             display="硅胶LED UV喷"),
    dict(key="silicone_water_base_transparent",   display="硅胶水基透明"),
    # --- Silicone (mostly template/empty) ---
    dict(key="silicone_coarse_uv",                display="硅胶粗UV"),
    dict(key="silicone_sandblast_uv",             display="硅胶喷砂UV"),
    dict(key="silicone_wrinkle_uv",               display="硅胶皱纹UV"),
    dict(key="silicone_orange_peel_uv",           display="硅胶橘皮UV"),
    dict(key="silicone_watercolor_ink",           display="硅胶水彩油墨"),
    dict(key="silicone_mica_pearl",               display="硅胶珠光云母"),
    # --- UV Ink ---
    dict(key="uv_oil_matte",                      display="UV油哑油"),
    dict(key="water_varnish_matte",               display="水油哑油"),
    # --- Lamination ---
    dict(key="lamination_material",               display="过胶材料"),
    # --- Yaguang UV Oil ---
    dict(key="yaguang_uv_oil",                    display="压光UV油"),
]


def product_table_name(mod_key: str) -> str:
    """Return the product table name for a module key."""
    if mod_key == "lamination_material":
        return "lamination_material_products"
    return f"{mod_key}_product"


def compat_table_name(mod_key: str) -> str:
    """Return the compatibility table name for a module key."""
    if mod_key == "lamination_material":
        return "lamination_material_compatibility"
    return f"{mod_key}_compatibility"


def usage_column_name(mod_key: str) -> str:
    """Return the usage column name for a module key."""
    if mod_key == "lamination_material":
        return "usage_text"
    return "usage"


# Product column mapping is built dynamically in product_columns_for().

# Build per-module column list with correct usage column name.
# Also filters to only columns that actually exist in the table.
def product_columns_for(mod_key: str, p_tbl: str) -> list[tuple[str, str]]:
    """Return list of (db_column, chinese_header) for the product table.

    Queries information_schema to find which of the standard columns
    actually exist, using the correct usage column name per module.
    """
    usage_col = usage_column_name(mod_key)

    # Define our desired columns with their Chinese headers
    desired = [
        ("material_code", "物料编号"),
        ("material_name", "物料名称"),
        (usage_col, "用途"),
        ("category", "性质"),
        ("color", "颜色"),
        ("responsible_person", "测试员"),
        ("stock_code", "库存编号"),
        ("supplier_code", "供应商编号"),
    ]

    # Query which columns actually exist in this table
    safe_tbl = _sanitize_tbl(p_tbl)
    existing = set()
    col_rows = query(
        f"SELECT column_name FROM information_schema.columns "
        f"WHERE table_schema='public' AND table_name='{safe_tbl}'"
    )
    for r in col_rows:
        existing.add(r.get("column_name", ""))

    # Filter desired columns to only those that exist
    result = []
    for db_col, cn_header in desired:
        if db_col in existing:
            result.append((db_col, cn_header))

    return result


# ---------------------------------------------------------------------------
# Helper: DB connection and query utilities
# ---------------------------------------------------------------------------

def _sanitize_tbl(name: str) -> str:
    """Strip to safe SQL identifier (alphanum + underscore only)."""
    return "".join(c for c in name if c.isalnum() or c == "_")


def query(sql: str):
    """Execute SQL via psql and return a list of dicts.

    NOTE: We read raw bytes and decode as utf-8 to avoid
    Windows GBK encoding issues with Chinese characters
    in database output.
    """
    import subprocess
    import csv
    import io

    env = os.environ.copy()
    env["PGPASSWORD"] = DB_CONFIG["password"]
    env["PGCLIENTENCODING"] = "UTF8"

    cmd = [
        "psql",
        "-h", DB_CONFIG["host"],
        "-p", str(DB_CONFIG["port"]),
        "-U", DB_CONFIG["user"],
        "-d", DB_CONFIG["dbname"],
        "-A",                      # unaligned output
        "-F", "\t",                # tab separator
        "--no-psqlrc",
        "-c", sql,
    ]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,   # raw bytes, no text=True
            timeout=30,
            env=env,
        )
    except subprocess.TimeoutExpired:
        print(f"  [WARN] SQL timed out: {sql[:80]}...")
        return []
    except FileNotFoundError:
        print("  [ERROR] psql not found on PATH. Install PostgreSQL client.")
        sys.exit(1)

    if result.returncode != 0:
        stderr = result.stderr
        if stderr:
            try:
                msg = stderr.decode("utf-8", errors="replace").strip()
            except Exception:
                msg = str(stderr).strip()
            if msg:
                print(f"  [WARN] psql error (exit={result.returncode})")
        return []

    try:
        raw = result.stdout.decode("utf-8", errors="replace").strip()
    except Exception:
        raw = ""
    if not raw:
        return []

    # psql with -A -F'\t' outputs tab-separated data.
    # The first line is the header row (column names), subsequent lines are data.
    lines = raw.split("\n")
    if len(lines) < 2:
        return []

    headers = lines[0].strip().split("\t")
    data_lines = lines[1:]  # skip header

    rows = []
    for line in data_lines:
        line = line.strip()
        if not line:
            continue
        vals = line.split("\t")
        row = {}
        for i, h in enumerate(headers):
            row[h] = vals[i] if i < len(vals) else ""
        rows.append(row)

    return rows


def table_exists(tbl: str) -> bool:
    """Check if a table exists in the public schema."""
    # Sanitize: only allow alphanumeric and underscore
    safe_tbl = _sanitize_tbl(tbl)
    rows = query(
        f"SELECT 1 FROM information_schema.tables "
        f"WHERE table_schema='public' AND table_name='{safe_tbl}'"
    )
    return len(rows) > 0


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _write_product_sheet(ws, prod_rows: list[dict], mod_key: str, p_tbl: str) -> int:
    """Write product data into ws (Sheet 1). Returns row count (excluding header)."""
    cols = product_columns_for(mod_key, p_tbl)
    headers = [h for _, h in cols]
    db_cols = [c for c, _ in cols]

    # Header
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    if not prod_rows:
        return 0

    # Data
    for ri, row in enumerate(prod_rows, 2):
        for ci, db_col in enumerate(db_cols, 1):
            val = row.get(db_col, "") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    return len(prod_rows)


def _step_column_name(mod_key: str) -> str:
    """Return the step column name for a compatibility table."""
    if mod_key == "yaguang_uv_oil":
        return "step_name"
    return "post_processing_step"


def _write_compat_sheet(ws, prod_rows: list[dict], compat_rows: list[dict],
                         mod_key: str, prod_count: int) -> int:
    """Write the compatibility matrix (Sheet 2) in wide-format pivot.

    Returns the number of rows written (excluding header).
    If no products, writes a note and returns 0.
    """
    if prod_count == 0 or not compat_rows:
        ws.cell(row=1, column=1, value="兼容性矩阵").font = HEADER_FONT
        ws.cell(row=2, column=1, value="该模块无产品数据，兼容性矩阵为空。")
        return 0

    step_col = _step_column_name(mod_key)

    # Build lookup: product_id -> {step_name: status, ...}
    # Also collect all unique step names preserving order
    compat_map = {}   # product_id -> OrderedDict
    all_steps = OrderedDict()
    # Also need product_id -> material_code/name
    prod_lookup = {}
    for r in prod_rows:
        pid = str(r.get("id", "")).strip()
        prod_lookup[pid] = r

    for r in compat_rows:
        pid = str(r.get("product_id", "")).strip()
        step = (r.get(step_col, "") or "").strip()
        status = (r.get("compatibility_status", "") or "").strip()
        if pid not in compat_map:
            compat_map[pid] = OrderedDict()
        compat_map[pid][step] = status
        if step not in all_steps:
            all_steps[step] = None

    step_list = list(all_steps.keys())

    # Header row
    fixed_headers = ["物料编号", "物料名称"]
    headers = fixed_headers + step_list
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows — one per product that appears in compat_map
    row_idx = 2
    for pid in compat_map:
        prod = prod_lookup.get(pid, {})
        mat_code = prod.get("material_code", "") or ""
        mat_name = prod.get("material_name", "") or ""
        ws.cell(row=row_idx, column=1, value=mat_code).font = BODY_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=mat_name).font = BODY_FONT
        ws.cell(row=row_idx, column=2).border = THIN_BORDER

        status_map = compat_map[pid]
        for si, step in enumerate(step_list):
            col = 3 + si
            val = status_map.get(step, "")
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            if val == "V":
                cell.fill = GREEN_FILL
            elif val == "X":
                cell.fill = RED_FILL
        row_idx += 1

    compat_row_count = row_idx - 2  # exclude header

    # Freeze top row and first 2 columns
    ws.freeze_panes = "C2"

    # Auto-fit column widths (approximate)
    _auto_fit_columns(ws, headers)

    return compat_row_count


def _auto_fit_columns(ws, headers):
    """Approximate column width auto-fitting."""
    for ci, h in enumerate(headers, 1):
        # Estimate width: max of header length and data content
        max_len = len(h.encode("utf-8"))  # rough Chinese char width
        for row in ws.iter_rows(min_col=ci, max_col=ci, min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    # CJK characters are ~2x width
                    val = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                    if char_len > max_len:
                        max_len = char_len
        # Cap at 60 and minimum 8
        width = min(max(max_len * 0.85, 8), 60)
        ws.column_dimensions[get_column_letter(ci)].width = width


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------
def main():
    out_dir = "F:/YTS+JYY/审核导出"
    os.makedirs(out_dir, exist_ok=True)

    summary_rows = []
    all_ok = True

    for mod in MODULES:
        key = mod["key"]
        display = mod["display"]
        p_tbl = product_table_name(key)
        c_tbl = compat_table_name(key)

        print(f"\n[{display}] ({key})")
        print(f"  产品表: {p_tbl}  |  兼容表: {c_tbl}")

        # Check tables exist
        p_exists = table_exists(p_tbl)
        c_exists = table_exists(c_tbl)

        if not p_exists:
            print(f"  [SKIP] 产品表 {p_tbl} 不存在")
            summary_rows.append((display, "—", "—", "表不存在"))
            continue

        # Fetch products
        prod_cols_list = product_columns_for(key, p_tbl)
        prod_cols = ["id"] + [c for c, _ in prod_cols_list]
        prod_sql_cols = ", ".join(prod_cols)

        prod_rows = query(f"SELECT {prod_sql_cols} FROM {_sanitize_tbl(p_tbl)} ORDER BY id")
        prod_count = len(prod_rows)
        print(f"  产品数量: {prod_count}")

        # Fetch compatibility
        compat_rows = []
        compat_count = 0
        if c_exists and prod_count > 0:
            step_col = _step_column_name(key)
            # Check if display_order column exists
            has_display_order = False
            do_rows = query(
                f"SELECT column_name FROM information_schema.columns "
                f"WHERE table_schema='public' AND table_name='{_sanitize_tbl(c_tbl)}' "
                f"AND column_name='display_order'"
            )
            has_display_order = len(do_rows) > 0
            order_clause = f"product_id, {step_col}"
            if has_display_order:
                order_clause = f"product_id, display_order, {step_col}"
            compat_cols = f"product_id, {step_col}, compatibility_status"
            compat_rows = query(
                f"SELECT {compat_cols} FROM {_sanitize_tbl(c_tbl)} ORDER BY {order_clause}"
            )
            compat_count = len(compat_rows)
            print(f"  兼容记录: {compat_count}")

        # Create workbook
        wb = openpyxl.Workbook()

        # Sheet 1: 产品列表
        ws1 = wb.active
        ws1.title = "产品列表"
        actual_prod_count = _write_product_sheet(ws1, prod_rows, key, p_tbl)
        ws1.freeze_panes = "A2"  # freeze header row

        # Sheet 2: 兼容性矩阵
        ws2 = wb.create_sheet("兼容性矩阵")
        actual_compat_count = _write_compat_sheet(ws2, prod_rows, compat_rows, key, prod_count)

        # Save
        fname = f"{key}_审核.xlsx"
        fpath = os.path.join(out_dir, fname)
        wb.save(fpath)
        print(f"  [OK] 已保存: {fpath}")

        # Status
        status = "OK"
        if prod_count == 0:
            status = "(空)"
        summary_rows.append((display, prod_count, compat_count, status))

    # -----------------------------------------------------------------------
    # Summary workbook
    # -----------------------------------------------------------------------
    print("\n\n===== 生成汇总清单 =====")
    wb_summary = openpyxl.Workbook()
    ws = wb_summary.active
    ws.title = "汇总清单"

    sum_headers = ["模块名称", "产品数量", "兼容记录数", "状态"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    total_prod = 0
    total_compat = 0
    ok_count = 0
    empty_count = 0
    skip_count = 0
    for ri, (disp, pc, cc, st) in enumerate(summary_rows, 2):
        ws.cell(row=ri, column=1, value=disp).border = THIN_BORDER
        ws.cell(row=ri, column=1).font = BODY_FONT
        ws.cell(row=ri, column=2, value=str(pc) if pc != "—" else "—").border = THIN_BORDER
        ws.cell(row=ri, column=2).font = BODY_FONT
        ws.cell(row=ri, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=ri, column=3, value=str(cc) if cc != "—" else "—").border = THIN_BORDER
        ws.cell(row=ri, column=3).font = BODY_FONT
        ws.cell(row=ri, column=3).alignment = Alignment(horizontal="center")

        status_cell = ws.cell(row=ri, column=4, value=st)
        status_cell.border = THIN_BORDER
        status_cell.font = BODY_FONT
        status_cell.alignment = Alignment(horizontal="center")
        if st == "OK":
            ok_count += 1
        elif st == "(空)":
            empty_count += 1
        elif st == "表不存在":
            skip_count += 1
        if isinstance(pc, int):
            total_prod += pc
        if isinstance(cc, int):
            total_compat += cc

    # Summary row
    sum_row = len(summary_rows) + 2
    ws.cell(row=sum_row, column=1, value="合计").font = Font(bold=True, size=11)
    ws.cell(row=sum_row, column=1).border = THIN_BORDER
    ws.cell(row=sum_row, column=2, value=total_prod).font = Font(bold=True, size=11)
    ws.cell(row=sum_row, column=2).border = THIN_BORDER
    ws.cell(row=sum_row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=sum_row, column=3, value=total_compat).font = Font(bold=True, size=11)
    ws.cell(row=sum_row, column=3).border = THIN_BORDER
    ws.cell(row=sum_row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=sum_row, column=4, value=f"OK={ok_count}, 空={empty_count}, 跳过={skip_count}").font = Font(bold=True, size=11)
    ws.cell(row=sum_row, column=4).border = THIN_BORDER

    _auto_fit_columns(ws, sum_headers)
    ws.freeze_panes = "A2"

    sum_path = os.path.join(out_dir, "00_汇总清单.xlsx")
    wb_summary.save(sum_path)
    print(f"[OK] 汇总清单: {sum_path}")

    # -----------------------------------------------------------------------
    # Validation report (Markdown)
    # -----------------------------------------------------------------------
    print("\n===== 验证报告 =====")
    report_lines = [
        f"# 审核导出报告",
        f"",
        f"**导出时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"",
        f"## 模块清单",
        f"",
        f"| 模块名称 | 产品数量 | 兼容记录数 | 状态 | 文件名 |",
        f"|---------|---------|-----------|------|--------|",
    ]

    report_all_ok = True
    for mod in MODULES:
        key = mod["key"]
        fname = f"{key}_审核.xlsx"
        fpath = os.path.join(out_dir, fname)
        exists = os.path.isfile(fpath)

        # Find the summary entry
        entry = next(
            (s for s in summary_rows if s[0] == mod["display"]),
            (mod["display"], "—", "—", "未找到"),
        )
        disp, pc, cc, st = entry

        if not exists:
            st = "文件缺失"
            report_all_ok = False

        report_lines.append(
            f"| {disp} | {pc} | {cc} | {st} | {fname} |"
        )

    report_lines.extend([
        "",
        "---",
        "",
        "## 汇总",
        f"**总模块数**: {len(MODULES)}",
        f"**总产品数**: {total_prod}",
        f"**总兼容记录数**: {total_compat}",
        f"**正常模块**: {ok_count}",
        f"**空模块**: {empty_count}",
        f"**跳过(表不存在)**: {skip_count}",
        "",
        f"**整体状态**: {'全部通过' if report_all_ok else '存在问题，请检查上表'}",
    ])

    report_path = os.path.join(out_dir, "00_审核导出报告.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))
    print(f"[OK] 导出报告: {report_path}")

    # Final status
    print(f"\n{'='*60}")
    print(f"导出完成! 文件位于: {out_dir}")
    ok_count_excel = sum(1 for m in MODULES if os.path.isfile(os.path.join(out_dir, f"{m['key']}_审核.xlsx")))
    print(f"  模块文件: {ok_count_excel}/{len(MODULES)}")
    print(f"  汇总清单: {sum_path}")
    print(f"  导出报告: {report_path}")
    print(f"{'='*60}")

    return 0 if report_all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
