#!/usr/bin/env python3
"""分析 LEO 纸品 xlsx 文件结构"""

import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.utils import get_column_letter

base = "G:/雅图仕/download/LEO纸品"
files = [
    f"{base}/书板贴纸-应用指引书-20251201 (2).xlsx",
    f"{base}/平面产品贴纸-应用指引书-20250701 (7).xlsx",
    f"{base}/非平面产品贴纸-应用指引书-20250701 (7).xlsx",
]


def sanitize(v):
    """安全转义，避免 GBK 编码问题"""
    if v is None:
        return ""
    s = str(v).strip().replace("\n", "\\n").replace("\r", "\\r")
    return s[:60]


def analyze_file(path):
    print(f"{'=' * 100}")
    print(f"文件: {path}")
    print(f"{'=' * 100}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  打开失败: {e}")
        return

    print(f"  Sheet 列表: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  --- Sheet: [{sn}] ---")
        print(f"    max_row={ws.max_row}, max_column={ws.max_column}")

        # 前 15 行的浓缩内容 (col 1-15)
        print(f"    --- 前15行, 列1-15 ---")
        for r in range(1, min(16, ws.max_row + 1)):
            cells = []
            for c in range(1, min(ws.max_column + 1, 16)):
                s = sanitize(ws.cell(r, c).value)
                if s:
                    cells.append(f"{get_column_letter(c)}={s}")
            if cells:
                print(f"    Row {r:3d}: {' | '.join(cells)}")
            else:
                print(f"    Row {r:3d}: (全部空白)")

        # 检查列40-130的内容
        print(f"    --- 前20行, 列40-130 非空样本 ---")
        nonempty_in_wide = 0
        for r in range(1, min(21, ws.max_row + 1)):
            for c in range(40, min(ws.max_column + 1, 131)):
                v = ws.cell(r, c).value
                if v is not None:
                    nonempty_in_wide += 1
                    if nonempty_in_wide <= 15:
                        s = sanitize(v)
                        print(f"      Row {r}, Col {c} ({get_column_letter(c)}): '{s}'")
        print(f"    第40-130列非空单元格总数(前20行): {nonempty_in_wide}")

        # 检查合并单元格
        if ws.merged_cells.ranges:
            print(f"    合并单元格 ({len(ws.merged_cells.ranges)} 个), 前15个:")
            for i, mc in enumerate(list(ws.merged_cells.ranges)[:15]):
                print(f"      {mc}")
        else:
            print(f"    合并单元格: 无")

        # 检查第6列(F)和第9列(I)的内容
        print(f"    --- 关键列检查 (F=col6, I=col9) ---")
        for r in range(1, min(ws.max_row + 1, 51)):
            c6s = sanitize(ws.cell(r, 6).value)
            c9s = sanitize(ws.cell(r, 9).value)
            if c6s or c9s:
                print(f"      Row {r:3d}: F='{c6s[:60]}'  |  I='{c9s[:60]}'")

    wb.close()
    print()


for f in files:
    analyze_file(f)
