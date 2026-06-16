#!/usr/bin/env python3
"""
P0 ETL Script — Parse 新物料適用性指引書 xlsx files from G: drive and import into PostgreSQL.

Usage:
    python p0_etl.py                          # Full import
    python p0_etl.py --dry-run                # Print what would be done
    python p0_etl.py --category silicone_glow_ink  # Single category
    python p0_etl.py --validate               # Run SELECT COUNT(*) on all tables
    python p0_etl.py --create-tables-only     # Only create tables, no import
"""

import argparse
import glob
import os
import re
import sys
import time
from collections import OrderedDict

import openpyxl

# ──────────────────────────────────────────────
# DB Configuration
# ──────────────────────────────────────────────
DB_HOST = "localhost"
DB_PORT = 5432
DB_NAME = "gold_foil_db"
DB_USER = "postgres"
DB_PASS = "HryENprJrxThYSDz"

BASE_DIR = r"G:/雅图仕/download"

# ──────────────────────────────────────────────
# CATEGORIES — 15 P0 categories (OrderedDict)
# Each entry:
#   key = short slug
#   value = {
#       dir: subdirectory under BASE_DIR,
#       glob: file pattern,
#       data_count: known rows (0 = empty, create empty tables),
#       product_table: PG product table name,
#       compat_table: PG compat table name,
#       schema_type: "silicone" | "yaguang" | "uv_oil_matte" | "water_varnish" | "lamination",
#       subtype: For schema_type=silicone, the specific file subtype pattern
#   }
# ──────────────────────────────────────────────
CATEGORIES = OrderedDict()

# --- Silicone (硅胶) — 9 subtypes, 7 with data ---
CATEGORIES["silicone_glow_ink"] = {
    "dir": "硅胶",
    "glob": "*发光油墨*",
    "data_count": 2,
    "product_table": "silicone_glow_ink_product",
    "compat_table": "silicone_glow_ink_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_white_uv"] = {
    "dir": "硅胶",
    "glob": "*白UV*",
    "data_count": 2,
    "product_table": "silicone_white_uv_product",
    "compat_table": "silicone_white_uv_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_glittering_star"] = {
    "dir": "硅胶",
    "glob": "*Glittering-Star*",
    "data_count": 7,
    "product_table": "silicone_glittering_star_product",
    "compat_table": "silicone_glittering_star_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_screen_uv"] = {
    "dir": "硅胶",
    "glob": "*丝印UV*",
    "data_count": 5,
    "product_table": "silicone_screen_uv_product",
    "compat_table": "silicone_screen_uv_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_led_uv_spray"] = {
    "dir": "硅胶",
    "glob": "*LED UV喷涂*",
    "data_count": 1,
    "product_table": "silicone_led_uv_spray_product",
    "compat_table": "silicone_led_uv_spray_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_water_base_transparent"] = {
    "dir": "硅胶",
    "glob": "*水性透明白墨*",
    "data_count": 1,
    "product_table": "silicone_water_base_transparent_product",
    "compat_table": "silicone_water_base_transparent_compatibility",
    "schema_type": "silicone",
}

# Silicone subtypes with NO data rows (need empty tables for CRUD)
CATEGORIES["silicone_coarse_uv"] = {
    "dir": "硅胶",
    "glob": "*粗纹UV*",
    "data_count": 0,
    "product_table": "silicone_coarse_uv_product",
    "compat_table": "silicone_coarse_uv_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_orange_peel_uv"] = {
    "dir": "硅胶",
    "glob": "*桔纹UV*",
    "data_count": 0,
    "product_table": "silicone_orange_peel_uv_product",
    "compat_table": "silicone_orange_peel_uv_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_sandblast_uv"] = {
    "dir": "硅胶",
    "glob": "*磨砂UV*",
    "data_count": 0,
    "product_table": "silicone_sandblast_uv_product",
    "compat_table": "silicone_sandblast_uv_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_wrinkle_uv"] = {
    "dir": "硅胶",
    "glob": "*皱纹UV*",
    "data_count": 0,
    "product_table": "silicone_wrinkle_uv_product",
    "compat_table": "silicone_wrinkle_uv_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_watercolor_ink"] = {
    "dir": "硅胶",
    "glob": "*水彩油墨*",
    "data_count": 0,
    "product_table": "silicone_watercolor_ink_product",
    "compat_table": "silicone_watercolor_ink_compatibility",
    "schema_type": "silicone",
}

CATEGORIES["silicone_mica_pearl"] = {
    "dir": "硅胶",
    "glob": "*Mica Pearl*",
    "data_count": 0,
    "product_table": "silicone_mica_pearl_product",
    "compat_table": "silicone_mica_pearl_compatibility",
    "schema_type": "silicone",
}

# --- UV Ink (UV油墨) ---
CATEGORIES["uv_oil_matte"] = {
    "dir": "UV油墨",
    "glob": "*哑光UV油*",
    "data_count": 12,
    "product_table": "uv_oil_matte_product",
    "compat_table": "uv_oil_matte_compatibility",
    "schema_type": "uv_oil_matte",
}

CATEGORIES["water_varnish_matte"] = {
    "dir": "UV油墨",
    "glob": "*哑光水油*",
    "data_count": 36,
    "product_table": "water_varnish_matte_product",
    "compat_table": "water_varnish_matte_compatibility",
    "schema_type": "water_varnish",
}

# --- Lamination (印刷加工) — reference table, NOT product+compat ---
CATEGORIES["lamination_material"] = {
    "dir": "印刷加工",
    "glob": "*后加工资料标准书*",
    "data_count": 24,
    "product_table": "lamination_material_products",
    "compat_table": "lamination_material_compatibility",
    "schema_type": "lamination",
}

# ──────────────────────────────────────────────
# DDL Templates
# ──────────────────────────────────────────────

# Shared product DDL — all product+compat types use the same column names
PRODUCT_DDL = """
CREATE TABLE IF NOT EXISTS {table} (
    id SERIAL PRIMARY KEY,
    material_code VARCHAR(100) DEFAULT NULL,
    supplier_code VARCHAR(100) DEFAULT NULL,
    stock_code VARCHAR(100) DEFAULT NULL,
    material_name VARCHAR(500) NOT NULL,
    usage TEXT DEFAULT NULL,
    category VARCHAR(200) DEFAULT NULL,
    color VARCHAR(200) DEFAULT NULL,
    responsible_person VARCHAR(200) DEFAULT NULL,
    min_sheet_size VARCHAR(200) DEFAULT NULL,
    max_sheet_size VARCHAR(200) DEFAULT NULL,
    min_spacing VARCHAR(100) DEFAULT NULL,
    max_spacing VARCHAR(100) DEFAULT NULL,
    design_info TEXT DEFAULT NULL,
    applicable_interface TEXT DEFAULT NULL,
    notes TEXT DEFAULT NULL,
    is_active BOOLEAN DEFAULT TRUE,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
"""

COMPAT_DDL = """
CREATE TABLE IF NOT EXISTS {table} (
    id SERIAL PRIMARY KEY,
    product_id INTEGER NOT NULL REFERENCES {product_table}(id),
    post_processing_step VARCHAR(500) NOT NULL,
    compatibility_status VARCHAR(10) NOT NULL,
    remark TEXT DEFAULT NULL,
    display_order INTEGER DEFAULT 0,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
"""

LAMINATION_TABLE_DDL = """
CREATE TABLE IF NOT EXISTS lamination_material_products (
    id SERIAL PRIMARY KEY,
    material_code VARCHAR(100) NOT NULL,
    stock_code VARCHAR(200) DEFAULT NULL,
    material_name VARCHAR(500) NOT NULL,
    usage_text TEXT DEFAULT NULL,
    material_type VARCHAR(200) DEFAULT NULL,
    thickness_film VARCHAR(100) DEFAULT NULL,
    thickness_glue VARCHAR(100) DEFAULT NULL,
    size_info VARCHAR(200) DEFAULT NULL,
    color VARCHAR(200) DEFAULT NULL,
    shape VARCHAR(200) DEFAULT NULL,
    responsible_person VARCHAR(200) DEFAULT NULL,
    category VARCHAR(200) DEFAULT NULL,
    is_active BOOLEAN DEFAULT TRUE,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS lamination_material_compatibility (
    id SERIAL PRIMARY KEY,
    product_id INTEGER NOT NULL REFERENCES lamination_material_products(id),
    post_processing_step VARCHAR(500) NOT NULL,
    compatibility_status VARCHAR(10) NOT NULL,
    remark TEXT DEFAULT NULL,
    display_order INTEGER DEFAULT 0,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
"""

# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")


def normalize_multiline(v):
    """Replace \\n separators with ' / ' in cell values, strip whitespace."""
    if v is None:
        return ""
    s = str(v).strip()
    # Replace escaped newlines (literal \\n in some odd xlsx)
    s = s.replace("\\n", " / ")
    # Replace actual newlines
    s = s.replace("\n", " / ")
    # Collapse multiple separators
    while " /  / " in s:
        s = s.replace(" /  / ", " / ")
    while "  / " in s:
        s = s.replace("  / ", " / ")
    s = s.strip(" / ")
    return s.strip()


def cell_text(ws, row, col):
    """Safely get cell text, handling merged cells that return None on non-top-left."""
    try:
        v = ws.cell(row, col).value
    except Exception:
        return ""
    if v is None:
        return ""
    return normalize_multiline(v)


def is_valid_status(v):
    """Check if a cell value is a real compatibility status (V/X/▷/O), not descriptive text."""
    s = str(v).strip()
    return s in ("V", "X", "▷", "O")


# ──────────────────────────────────────────────
# Sheet Parsing
# ──────────────────────────────────────────────

def find_sheet(ws):
    """
    Identify the correct sheet in the workbook.
    Try the main overview sheet first (not version history, not reference).
    """
    # Prefer the first sheet that has "適用性指引書" in name
    for sn in ws.parent.sheetnames:
        if "適用性" in sn or "指引" in sn:
            return ws.parent[sn]

    # Fallback: first non-version sheet with enough rows
    for sn in ws.parent.sheetnames:
        sheet = ws.parent[sn]
        if sheet.max_row > 15 and sheet.max_column > 60:
            return sheet

    return ws


def find_step_header_row(ws):
    """
    Find the row containing step header names (the row with step names
    like 單面印刷, 雙面印刷, etc. starting around col 40-41).

    The step header is identified by having >= 20 non-empty cells in cols 40-60
    and containing known step keywords.

    Returns:
        int: The row number of the step-level header (third row in 3-level header).
    """
    step_keywords = {"單面印刷", "雙面印刷", "連線印刷", "離線印刷", "油性油",
                     "水性油", "UV油", "普通油", "普通燙金", "鐳射燙金",
                     "立體燙金", "擊凸", "擊凹", "過膠", "絲印", "植毛",
                     "機啤", "手啤", "壓紋", "粘貼", "車線"}

    best_row = None
    best_count = 0

    for row in range(15, min(ws.max_row + 1, 30)):
        count = 0
        for col in range(40, min(ws.max_column + 1, 120)):
            v = cell_text(ws, row, col)
            if not v:
                continue
            # Check if this looks like a step name
            if any(kw in v for kw in step_keywords) or (len(v) >= 2 and len(v) <= 20):
                count += 1

        if count > best_count:
            best_count = count
            best_row = row

    return best_row


def find_category_header_row(ws):
    """
    Find the row containing category group headers (印刷, 燙金, 過膠, 絲印, etc.
    typically one row above the step header).

    Returns:
        int: Row number for category group, or None.
    """
    step_row = find_step_header_row(ws)
    if step_row is None:
        return None

    # The category row is usually 1-2 rows above the step row
    group_keywords = {"印刷", "燙金", "過膠", "絲印", "植毛", "啤", "手工", "其他",
                      "物料", "適用", "不適用"}

    for row in range(max(14, step_row - 5), step_row):
        count = 0
        for col in range(40, min(ws.max_column + 1, 120)):
            v = cell_text(ws, row, col)
            if not v:
                continue
            if any(kw == v for kw in group_keywords) or (len(v) <= 4 and len(v) >= 1):
                count += 1
        if count >= 3:
            return row

    return step_row - 1 if step_row else None


def find_data_rows(ws, step_row):
    """
    Find all data rows in the sheet.

    A data row must:
    - Have non-empty col6 (testing code)
    - Col9 (material name) must be non-empty
    - Col6 starts with '#' (primary) or matches pattern like UV-OS-XXXXX / OS-XXXXX / CS-XXXXX
    - Skip rows with '/' alone in col6
    - Skip header/metadata rows (rows above step_row, rows with fewer cols populated)

    Returns:
        list of dict: Each has {row, col6_value, is_primary}
    """
    data_rows = []
    seen_codes = set()

    for row in range(step_row + 1, ws.max_row + 1):
        col6 = cell_text(ws, row, 6)
        col9 = cell_text(ws, row, 9)

        # Must have col9 non-empty (material name)
        if not col9 or col9 in ("/", "-"):
            continue

        # Must have col6 with a code
        if not col6 or col6 == "/" or col6 == "-":
            continue

        # Skip rows that look like documentation/notes (very long text in col6)
        if len(col6) > 200:
            continue

        # Check if col6 starts with # (primary) or is a valid code pattern
        is_primary = col6.startswith("#")
        code_pattern = re.match(
            r'^#?[A-Za-z]+-\d+|^[A-Za-z]+-\d+|^[A-Za-z]+\d+', col6
        )

        if not is_primary and not code_pattern:
            continue

        # Dedup
        first_code = col6.split(" / ")[0].strip().lstrip("#")
        if first_code in seen_codes:
            continue
        seen_codes.add(first_code)

        data_rows.append({
            "row": row,
            "code": col6,
            "name": col9,
            "is_primary": is_primary,
        })

    return data_rows


def extract_product_from_row(ws, row_num, schema_type="silicone"):
    """
    Extract product data from a row.

    Column mapping (1-indexed):
        C6  = material_code (NTD testing code)
        C7  = supplier_code (采购部申请编号)
        C8  = stock_code (物料型号/编号)
        C9  = material_name
        C10 = usage_text (物料用途)
        C11 = material_type/材质
        C12 = color (颜色) or thickness (厚度) for laminate
        C13 = responsible_person (测试员)
        C14 = color/尺寸 for some files

    Returns:
        dict of field->value, keys matching the product table columns
    """
    product = {
        "material_code": cell_text(ws, row_num, 6),
        "supplier_code": cell_text(ws, row_num, 7),
        "stock_code": cell_text(ws, row_num, 8),
        "material_name": cell_text(ws, row_num, 9),
        "usage": cell_text(ws, row_num, 10),
        "category": cell_text(ws, row_num, 11),
        "color": cell_text(ws, row_num, 14) if schema_type != "lamination" else None,
        "responsible_person": cell_text(ws, row_num, 13),
    }

    # For lamination, use different column mapping
    if schema_type == "lamination":
        product["usage_text"] = product.pop("usage")  # lamination table has usage_text
        product["material_type"] = cell_text(ws, row_num, 11)
        product["thickness_film"] = cell_text(ws, row_num, 12)
        product["thickness_glue"] = cell_text(ws, row_num, 13) if cell_text(ws, row_num, 13) else None
        product["size_info"] = cell_text(ws, row_num, 14)
        product["color"] = cell_text(ws, row_num, 15)
        product["shape"] = cell_text(ws, row_num, 16)
        product["category"] = None  # lamination uses material_type instead

    return product


def extract_compatibility_pairs(ws, row_num, step_row, cat_row=None):
    """
    Extract compatibility pairs from cols 40+ for one data row.

    Three-level header structure:
        Row cat_row  : group category (印刷, 烫金, etc.)
        Row step_row : individual step names

    Returns:
        list of (step_name, status) tuples, only for V/X/▷/O values.
    """
    pairs = []

    # Determine end column (either ws.max_column or 125 max)
    max_col = min(ws.max_column, 125)

    for col in range(40, max_col + 1):
        v = cell_text(ws, row_num, col)
        if not v:
            continue

        # Only V/X/▷/O count as real status values
        raw = v.strip()
        if raw in ("V", "X", "▷", "O"):
            # Get step name from step header row
            step_name = cell_text(ws, step_row, col)
            if step_name:
                # Optionally prepend category group
                if cat_row:
                    cat_name = cell_text(ws, cat_row, col)
                    if cat_name and cat_name != step_name:
                        # Check if cat_row value is the group label or an actual step
                        # Group labels are short (1-4 chars), step names are longer
                        if len(cat_name) <= 4 and cat_name in (
                            "印刷", "燙金", "過膠", "絲印", "植毛", "啤", "手工", "其他",
                            "適用", "不適用", "物料"
                        ):
                            full_name = f"{cat_name}/{step_name}"
                        else:
                            full_name = step_name
                    else:
                        full_name = step_name
                else:
                    full_name = step_name

                pairs.append((full_name, raw))

    return pairs


# ──────────────────────────────────────────────
# Database Operations
# ──────────────────────────────────────────────

def get_connection():
    """Get a psycopg2 database connection."""
    try:
        import psycopg2
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            dbname=DB_NAME,
            user=DB_USER,
            password=DB_PASS,
        )
        conn.autocommit = False
        return conn
    except ImportError:
        log("psycopg2 not available, trying psycopg2-binary...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "psycopg2-binary"])
        import psycopg2
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            dbname=DB_NAME,
            user=DB_USER,
            password=DB_PASS,
        )
        conn.autocommit = False
        return conn


def exec_sql(conn, sql, params=None):
    """Execute SQL with optional params."""
    with conn.cursor() as cur:
        if params:
            cur.execute(sql, params)
        else:
            cur.execute(sql)
        try:
            return cur.fetchall()
        except Exception:
            return []


def create_tables(conn, category_key, cat_info):
    """Create product and compatibility tables for a category."""
    schema_type = cat_info["schema_type"]
    product_table = cat_info["product_table"]
    compat_table = cat_info.get("compat_table")

    if schema_type == "lamination":
        # Lamination uses its own table structure
        log(f"  Creating lamination tables...")
        parts = [p.strip() for p in LAMINATION_TABLE_DDL.split(";") if p.strip()]
        for part in parts:
            exec_sql(conn, part)
        return

    sql = PRODUCT_DDL.format(table=product_table)
    exec_sql(conn, sql)
    log(f"  Created table {product_table}")

    if compat_table:
        sql = COMPAT_DDL.format(table=compat_table, product_table=product_table)
        exec_sql(conn, sql)
        log(f"  Created table {compat_table}")

    conn.commit()


def insert_product(conn, table, product_data):
    """Insert a product row. Returns the new id or None."""
    # Build dynamic INSERT
    fields = []
    placeholders = []
    values = []

    for k, v in product_data.items():
        if v is not None and v != "" and v != "/":
            fields.append(k)
            placeholders.append("%s")
            values.append(v)

    if not fields:
        return None

    # Check for duplicate by material_code
    code_val = product_data.get("material_code", "")
    first_code = code_val.split(" / ")[0].strip()
    if first_code:
        existing = exec_sql(
            conn,
            f"SELECT id FROM {table} WHERE material_code LIKE %s",
            (f"{first_code}%",),
        )
        if existing:
            log(f"    Duplicate material_code '{first_code}', skipping")
            return None

    sql = (
        f"INSERT INTO {table} ({', '.join(fields)}) "
        f"VALUES ({', '.join(placeholders)}) RETURNING id"
    )
    try:
        result = exec_sql(conn, sql, values)
        if result:
            return result[0][0]
    except Exception as e:
        # Try without RETURNING (older PG or different schema)
        try:
            sql_no_return = (
                f"INSERT INTO {table} ({', '.join(fields)}) "
                f"VALUES ({', '.join(placeholders)})"
            )
            exec_sql(conn, sql_no_return, values)
            # Get last inserted id
            res = exec_sql(conn, "SELECT LASTVAL()")
            if res:
                return res[0][0]
        except Exception as e2:
            log(f"    Insert failed: {e2}")
            return None
    return None


def insert_compatibility(conn, table, product_id, step_name, status):
    """Insert a compatibility row. Skip if duplicate."""
    try:
        exec_sql(
            conn,
            f"INSERT INTO {table} (product_id, post_processing_step, compatibility_status) "
            f"VALUES (%s, %s, %s)",
            (product_id, step_name, status),
        )
        return True
    except Exception:
        return False


def insert_lamination_product(conn, product_data):
    """Insert a lamination product."""
    return insert_product(conn, "lamination_material_products", product_data)


def insert_lamination_compat(conn, product_id, step_name, status):
    """Insert a lamination compatibility row."""
    return insert_compatibility(conn, "lamination_material_compatibility", product_id, step_name, status)


# ──────────────────────────────────────────────
# File Processing
# ──────────────────────────────────────────────

def find_files(cat_info):
    """Find files matching the category glob pattern."""
    dir_path = os.path.join(BASE_DIR, cat_info["dir"])
    if not os.path.isdir(dir_path):
        log(f"  Directory not found: {dir_path}")
        return []

    pattern = os.path.join(dir_path, cat_info["glob"])
    files = glob.glob(pattern)

    # Filter out temp files (starting with ~$)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]

    return sorted(files)


def get_schema_type_for_category(cat_info):
    """Determine schema type from category info."""
    return cat_info.get("schema_type", "silicone")


def process_category(conn, category_key, cat_info, dry_run=False):
    """Process all files for one category."""
    log(f"\n{'='*60}")
    log(f"Category: {category_key}")
    log(f"{'='*60}")

    files = find_files(cat_info)
    if not files:
        log(f"  No matching files found (glob: {cat_info['glob']} in {cat_info['dir']})")
        return {"files": 0, "products": 0, "compat": 0, "duplicates": 0, "skipped": 0}

    log(f"  Found {len(files)} file(s): {[os.path.basename(f) for f in files]}")

    stats = {"files": 0, "products": 0, "compat": 0, "duplicates": 0, "skipped": 0}
    schema_type = get_schema_type_for_category(cat_info)

    for filepath in files:
        stats["files"] += 1
        fname = os.path.basename(filepath)
        log(f"\n  --- Processing: {fname} ---")

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb[wb.sheetnames[0]]
            ws = find_sheet(ws)
        except Exception as e:
            log(f"    Error opening file: {e}")
            stats["skipped"] += 1
            continue

        # Find step header row
        step_row = find_step_header_row(ws)
        if step_row is None:
            log(f"    Could not find step header row, skipping")
            stats["skipped"] += 1
            continue

        cat_row = find_category_header_row(ws)
        log(f"    Step header at row {step_row}, category header at row {cat_row}")

        # Find data rows
        data_rows = find_data_rows(ws, step_row)
        log(f"    Found {len(data_rows)} data rows")

        if not data_rows:
            continue

        # For lamination, process differently
        if schema_type == "lamination":
            for dr in data_rows:
                row = dr["row"]
                product = extract_product_from_row(ws, row, schema_type="lamination")

                if dry_run:
                    log(f"    [DRY-RUN] Row {row}: {product['material_code']} -> {product['material_name']}")
                    continue

                pid = insert_lamination_product(conn, product)
                if pid:
                    stats["products"] += 1

                    # Extract compatibility pairs
                    pairs = extract_compatibility_pairs(ws, row, step_row, cat_row)
                    for step_name, status in pairs:
                        insert_lamination_compat(conn, pid, step_name, status)
                        stats["compat"] += 1

            conn.commit()
            continue

        # For product+compat types (silicone, uv, water varnish)
        for dr in data_rows:
            row = dr["row"]
            product = extract_product_from_row(ws, row, schema_type)
            product["material_code"] = dr["code"]

            if dry_run:
                log(f"    [DRY-RUN] Row {row}: {dr['code'][:40]} -> {product['material_name']}")
                stats["products"] += 1
                pairs = extract_compatibility_pairs(ws, row, step_row, cat_row)
                log(f"      {len(pairs)} compatibility entries")
                stats["compat"] += len(pairs)
                continue

            pid = insert_product(conn, cat_info["product_table"], product)
            if pid:
                stats["products"] += 1

                pairs = extract_compatibility_pairs(ws, row, step_row, cat_row)
                for step_name, status in pairs:
                    insert_compatibility(conn, cat_info["compat_table"], pid, step_name, status)
                    stats["compat"] += 1
            else:
                stats["duplicates"] += 1

        try:
            conn.commit()
        except Exception as e:
            log(f"    Commit error: {e}")
            conn.rollback()

        wb.close()

    log(f"\n  Results: {stats['products']} products, {stats['compat']} compat pairs, "
        f"{stats['duplicates']} duplicates, {stats['skipped']} skipped")
    return stats


def validate_tables(conn):
    """Run SELECT COUNT(*) on all P0 tables."""
    log("\n" + "=" * 60)
    log("Validation: SELECT COUNT(*) on all P0 tables")
    log("=" * 60)

    # Collect all table names
    tables = set()
    for key, info in CATEGORIES.items():
        tables.add(info["product_table"])
        if info.get("compat_table"):
            tables.add(info["compat_table"])

    # Rollback first to clear any stale transaction state
    conn.rollback()

    for table in sorted(tables):
        try:
            result = exec_sql(conn, f"SELECT COUNT(*) FROM {table}")
            count = result[0][0] if result else 0
            log(f"  {table}: {count} rows")
        except Exception as e:
            log(f"  {table}: ERROR - {e}")
            conn.rollback()  # Clear error for next query


def create_tables_only(conn):
    """Create all tables without importing data."""
    log(f"\n{'='*60}")
    log("Creating all tables (no import)")
    log(f"{'='*60}")

    for key, info in CATEGORIES.items():
        log(f"\nCategory: {key}")
        create_tables(conn, key, info)


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="P0 ETL — Import xlsx files to PostgreSQL")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be done without importing")
    parser.add_argument("--category", type=str, default=None, help="Only process one category (slug name)")
    parser.add_argument("--validate", action="store_true", help="Run SELECT COUNT(*) on all tables")
    parser.add_argument("--create-tables-only", action="store_true", help="Create tables without importing data")
    args = parser.parse_args()

    log(f"Starting P0 ETL")
    log(f"  DB: {DB_HOST}:{DB_PORT}/{DB_NAME}")
    log(f"  Source: {BASE_DIR}")

    if args.validate:
        conn = get_connection()
        try:
            validate_tables(conn)
        finally:
            conn.close()
        return

    if args.create_tables_only:
        conn = get_connection()
        try:
            create_tables_only(conn)
            conn.commit()
        except Exception as e:
            log(f"Error creating tables: {e}")
            conn.rollback()
        finally:
            conn.close()
        return

    # Main import flow
    tot_products = 0
    tot_compat = 0
    tot_duplicates = 0
    tot_skipped = 0
    tot_files = 0

    conn = get_connection()
    try:
        # Create all tables first
        log(f"\n{'='*60}")
        log("Phase 1: Creating tables")
        log(f"{'='*60}")
        create_tables_only(conn)

        log(f"\n{'='*60}")
        log("Phase 2: Importing data")
        log(f"{'='*60}")

        for key, info in CATEGORIES.items():
            if args.category and key != args.category:
                continue

            stats = process_category(conn, key, info, dry_run=args.dry_run)
            tot_files += stats["files"]
            tot_products += stats["products"]
            tot_compat += stats["compat"]
            tot_duplicates += stats["duplicates"]
            tot_skipped += stats["skipped"]

    except Exception as e:
        log(f"Fatal error: {e}")
        conn.rollback()
        raise
    finally:
        conn.close()

    log(f"\n{'='*60}")
    log("SUMMARY")
    log(f"{'='*60}")
    if args.dry_run:
        log(f"  DRY RUN — no data written")
    log(f"  Files processed: {tot_files}")
    log(f"  Products inserted: {tot_products}")
    log(f"  Compatibility pairs: {tot_compat}")
    log(f"  Duplicates skipped: {tot_duplicates}")
    log(f"  Files skipped: {tot_skipped}")
    log(f"Done.")


# ──────────────────────────────────────────────
# CLI entry
# ──────────────────────────────────────────────

if __name__ == "__main__":
    main()
