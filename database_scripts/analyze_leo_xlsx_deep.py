#!/usr/bin/env python3
"""深入分析 LEO xlsx 的表头结构和兼容性列范围"""

import io, sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.utils import get_column_letter

base = "G:/雅图仕/download/LEO纸品"
files = [
    ("书板贴纸", f"{base}/书板贴纸-应用指引书-20251201 (2).xlsx"),
    ("平面产品贴纸", f"{base}/平面产品贴纸-应用指引书-20250701 (7).xlsx"),
    ("非平面产品贴纸", f"{base}/非平面产品贴纸-应用指引书-20250701 (7).xlsx"),
]


def sanitize(v):
    if v is None:
        return ""
    return str(v).strip().replace("\n", "\\n")[:80]


for label, path in files:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"\n{'=' * 100}")
    print(f"=== {label} ===")
    print(f"{'=' * 100}")

    # 1. 找到表头所在行（有 NTD 测试单号/物料型号 的行）
    header_label_row = None
    for r in range(1, ws.max_row + 1):
        c9 = sanitize(ws.cell(r, 9).value)
        if "物料型號" in c9 or "物料型号" in c9:
            header_label_row = r
            print(f"\n[表头标签行] Row {r}: I列='{c9}'")
            # 打印该行所有非空列
            cols = []
            for c in range(1, min(ws.max_column + 1, 160)):
                v = sanitize(ws.cell(r, c).value)
                if v:
                    cols.append(f"{get_column_letter(c)}='{v}'")
            for item in cols:
                print(f"  {item}")
            break

    # 2. 找到分类组行（印刷、烫金、过胶 等大类）
    cat_row = None
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(40, min(ws.max_column + 1, 160)):
            s = sanitize(ws.cell(r, c).value)
            if s in ("印刷", "燙金", "過膠", "絲印", "植毛", "啤", "手工", "其他"):
                vals.append(f"{get_column_letter(c)}='{s}'")
        if vals:
            cat_row = r
            print(f"\n[分类组行] Row {r}:")
            for item in vals:
                print(f"  {item}")
            break

    # 3. 找到步骤/工序行（具体工艺名称）
    step_row = None
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(40, min(ws.max_column + 1, 160)):
            s = sanitize(ws.cell(r, c).value)
            # 跳过中文/数字的单字符单元格（可能是序号标记）
            if s and len(s) >= 2:
                vals.append(f"{get_column_letter(c)}='{s}'")
        if len(vals) >= 5:
            step_row = r
            print(f"\n[步骤行] Row {r}: 共 {len(vals)} 个非空列")
            # 只显示前20个
            for item in vals[:20]:
                print(f"  {item}")
            break

    # 4. 从步骤行+1开始检查数据行
    if step_row:
        print(f"\n[数据行] 从 Row {step_row + 1} 开始:")
        data_start = step_row + 1
        data_rows_found = 0
        for r in range(data_start, ws.max_row + 1):
            col9 = sanitize(ws.cell(r, 9).value)
            if col9 and col9 not in ("/", "-", "/-"):
                data_rows_found += 1
                # 检查该行的兼容性数据
                compat_cells = []
                for c in range(40, min(ws.max_column + 1, 160)):
                    v = sanitize(ws.cell(r, c).value)
                    if v in ("V", "X", "▷", "O"):
                        compat_cells.append(f"{get_column_letter(c)}={v}")
                if data_rows_found <= 3:
                    # 检查其他列
                    other_cols = []
                    for c in range(1, 40):
                        s = sanitize(ws.cell(r, c).value)
                        if s:
                            other_cols.append(f"{get_column_letter(c)}='{s}'")
                    print(
                        f"  Row {r}: F='{sanitize(ws.cell(r, 6).value)}' | "
                        f"I='{col9}' | "
                        f"兼容性={len(compat_cells)}个 | "
                        f"其他列: {' | '.join(other_cols[:8])}"
                    )
        print(f"  共 {data_rows_found} 个物料数据行")

    # 5. 检查兼容性列的范围（V/X/▷/O 分布）
    print(f"\n[兼容性值分布] 所有数据行:")
    v_count = x_count = d_count = o_count = 0
    compat_cols = set()
    for r in range(max(10, (step_row or 10) + 1), ws.max_row + 1):
        for c in range(40, min(ws.max_column + 1, 160)):
            v = sanitize(ws.cell(r, c).value)
            if v == "V":
                v_count += 1
                compat_cols.add(c)
            elif v == "X":
                x_count += 1
                compat_cols.add(c)
            elif v == "▷":
                d_count += 1
                compat_cols.add(c)
            elif v == "O":
                o_count += 1
                compat_cols.add(c)
    print(f"  V={v_count}, X={x_count}, ▷={d_count}, O={o_count}")
    if compat_cols:
        min_c = min(compat_cols)
        max_c = max(compat_cols)
        print(f"  兼容性列范围: {get_column_letter(min_c)}({min_c}) ~ {get_column_letter(max_c)}({max_c})")
        print(f"  共 {len(compat_cols)} 个有兼容性值的列")

    # 6. 检查合并单元格覆盖哪些区域
    print(f"\n[合并单元格关键区域]:")
    for mc in ws.merged_cells.ranges:
        mc_str = str(mc)
        # 只关注40列之后的合并
        if ":" in mc_str:
            parts = mc_str.split(":")
            if len(parts) == 2:
                from_col = openpyxl.utils.column_index_from_string(
                    "".join(c for c in parts[0] if c.isalpha())
                )
                if from_col >= 40:
                    print(f"  {mc}")

    wb.close()
