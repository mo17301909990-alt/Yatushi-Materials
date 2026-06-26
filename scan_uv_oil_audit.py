#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""针对UV油墨xlsx的详细审计补充脚本"""

import openpyxl
from openpyxl.utils import get_column_letter
import json

FILE = r"G:/雅图仕/download/UV油墨/哑光UV油标准书-20250730 (10)(1).xlsx"
wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb.active

print("=" * 80)
print("【补充审计】详细列映射分析")
print("=" * 80)

# 步骤行(row 23) - 兼容性子步骤名称
print("\n【步骤行 (Row 23) - 从列AO(41)开始】")
for col_idx in range(41, ws.max_column + 1):
    val = ws.cell(row=23, column=col_idx).value
    if val is not None:
        cl = get_column_letter(col_idx)
        cat22 = ws.cell(row=22, column=col_idx).value
        cat21 = ws.cell(row=21, column=col_idx).value
        cat20 = ws.cell(row=20, column=col_idx).value
        step_name = str(val).strip().replace('\n', ' / ')
        print(f"  {cl:4s} row23={step_name:30s} | row22={str(cat22 or '').strip()[:20]:20s} | row21={str(cat21 or '').strip()[:20]:20s}")

# 分类行 (row 22) - 兼容性大类名称
print("\n【分类行 (Row 22) - 兼容性大类】")
for col_idx in range(41, ws.max_column + 1):
    val = ws.cell(row=22, column=col_idx).value
    if val is not None:
        cl = get_column_letter(col_idx)
        print(f"  {cl:4s} {str(val).strip().replace(chr(10),' / ')[:50]}")

# 物料信息列(Row 22 - 列A至列P)
print("\n【物料信息列 (Row 22 - 列A~P)】")
for col_idx in range(1, 17):
    cl = get_column_letter(col_idx)
    val = ws.cell(row=22, column=col_idx).value
    print(f"  {cl:4s} {str(val or '').strip().replace(chr(10),' / ')[:60]}")

# 设计/结构信息列(Row 22 - 列Q~Z)
print("\n【设计/结构信息列 (Row 22 - 列Q~AG)】")
for col_idx in range(17, 34):
    cl = get_column_letter(col_idx)
    val22 = ws.cell(row=22, column=col_idx).value
    val23 = ws.cell(row=23, column=col_idx).value
    print(f"  {cl:4s} row22={str(val22 or '').strip()[:30]:30s} | row23={str(val23 or '').strip()[:30]:30s}")

# 基材/油墨/涂布信息(Row 22 - 列AH~AN)
print("\n【基材/油墨/涂布信息列 (Row 22 - 列AH~AN)】")
for col_idx in range(34, 41):
    cl = get_column_letter(col_idx)
    val22 = ws.cell(row=22, column=col_idx).value
    val23 = ws.cell(row=23, column=col_idx).value
    print(f"  {cl:4s} row22={str(val22 or '').strip()[:40]:40s} | row23={str(val23 or '').strip()[:40]:40s}")

# 备注列 DH
print("\n【备注列 (DH)】")
val = ws.cell(row=22, column=112).value
print(f"  DH row22={str(val or '').strip()[:60]}")

# D-生产商备注 (DI-DR)
print("\n【D-生产商备注 (DI-DR)】")
for col_idx in range(113, 123):
    cl = get_column_letter(col_idx)
    val22 = ws.cell(row=22, column=col_idx).value
    val23 = ws.cell(row=23, column=col_idx).value
    has_data = False
    for r in range(26, 38):
        v = ws.cell(r, col_idx).value
        if v is not None:
            has_data = True
            break
    print(f"  {cl:4s} header={str(val22 or '').strip()[:50]:50s} | has_data={has_data}")

# 检查列12(L) 规格 的数据
print("\n【列L(规格)数据样本】")
for r in range(24, 38):
    v = ws.cell(r, 12).value
    print(f"  Row {r}: {repr(v)}")

# 检查列O(形状)数据样本
print("\n【列O(形状)数据样本】")
for r in range(24, 38):
    v = ws.cell(r, 15).value
    print(f"  Row {r}: {repr(v)}")

# 检查列P(测试员)数据
print("\n【列P(测试员)数据样本】")
for r in range(24, 38):
    v = ws.cell(r, 16).value
    print(f"  Row {r}: {repr(v)}")

# 检查列Q-R(用纸尺寸)数据
print("\n【列Q-R(用纸尺寸)数据】")
for r in range(26, 38):
    q = ws.cell(r, 17).value
    ri = ws.cell(r, 18).value
    print(f"  Row {r}: Q={repr(q)}, R={repr(ri)}")

# 检查列S-T(點)数据
print("\n【列S-T(點)数据】")
for r in range(26, 38):
    s = ws.cell(r, 19).value
    t = ws.cell(r, 20).value
    print(f"  Row {r}: S={repr(s)}, T={repr(t)}")

# 检查列W-X(间距)数据
print("\n【列W-X(间距)数据】")
for r in range(26, 38):
    w = ws.cell(r, 23).value
    x = ws.cell(r, 24).value
    print(f"  Row {r}: W={repr(w)}, X={repr(x)}")

# 检查列AA(适用产品)数据
print("\n【列AA(适用产品)数据样本】")
for r in range(26, 38):
    v = ws.cell(r, 27).value
    print(f"  Row {r}: {repr(str(v)[:80] if v else v)}")

# 检查AB-AG列(结构信息V/X)数据
print("\n【列AB-AG(结构V/X)数据 - product 1】")
r = 26
for col_idx in range(28, 34):
    cl = get_column_letter(col_idx)
    v = ws.cell(r, col_idx).value
    step_name = ws.cell(23, col_idx).value
    print(f"  {cl:4s} step={str(step_name or '').strip()[:20]:20s} value={repr(v)}")

# 检查AH-AN列(基材/油墨/涂布)数据
print("\n【列AH-AN(基材/油墨/涂布)数据 - product 1】")
for col_idx in range(34, 41):
    cl = get_column_letter(col_idx)
    v = ws.cell(26, col_idx).value
    step_name = ws.cell(23, col_idx).value
    cat_name = ws.cell(22, col_idx).value
    print(f"  {cl:4s} row22={str(cat_name or '').strip()[:30]:30s} step={str(step_name or '').strip()[:20]:20s} value={repr(v)}")

# 步骤行(Row 22) 的兼容性大类列表
print("\n【兼容性大类汇总 (Row 22, AO~DH)】")
current_cat = None
for col_idx in range(41, ws.max_column + 1):
    v22 = ws.cell(row=22, column=col_idx).value
    if v22 is not None and str(v22).strip():
        current_cat = str(v22).strip().replace('\n', ' / ')
    cl = get_column_letter(col_idx)
    v23 = ws.cell(row=23, column=col_idx).value
    if v23 is not None and current_cat:
        print(f"  {cl:4s} cat={current_cat:15s} step={str(v23).strip().replace(chr(10),' / ')[:30]}")

wb.close()
