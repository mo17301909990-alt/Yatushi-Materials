#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""深度扫描 UV油墨 xlsx 文件结构"""

import openpyxl
from openpyxl.utils import get_column_letter
import json

FILE = r"G:/雅图仕/download/UV油墨/哑光UV油标准书-20250730 (10)(1).xlsx"

wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb.active

print("=" * 80)
print("【1. 文件基本信息】")
print("=" * 80)
print(f"Sheet 名称: {ws.title}")
print(f"最大行号: {ws.max_row}")
print(f"最大列号: {ws.max_column}")
print(f"数据区域: A1:{get_column_letter(ws.max_column)}{ws.max_row}")

# 合并单元格
print(f"\n【2. 合并单元格范围 ({len(ws.merged_cells.ranges)} 个)】")
for mc in sorted(ws.merged_cells.ranges, key=str):
    print(f"  {mc}")

# ====== 表头区域：第1-24行，逐列输出 ======
print(f"\n{'=' * 80}")
print(f"【3. 表头区域 (第1~24行) - 逐列输出】")
print(f"{'=' * 80}")

# 收集每列的表头信息
col_headers = {}
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    col_headers[col_letter] = []
    for row_idx in range(1, 25):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            col_headers[col_letter].append((row_idx, str(val).strip()))

print(f"\n列数: {len(col_headers)}")
for col_letter in sorted(col_headers.keys(), key=lambda x: (len(x), x)):
    entries = col_headers[col_letter]
    if entries:
        print(f"\n  列 {col_letter}:")
        for row_idx, val in entries:
            print(f"    第{row_idx}行: {repr(val)}")
    else:
        print(f"\n  列 {col_letter}: (空)")

# ====== 第24行以后的表头/数据行 ======
print(f"\n{'=' * 80}")
print(f"【4. 第24行精确解析 - 可能的真正表头】")
print(f"{'=' * 80}")
row24_cells = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=24, column=col_idx).value
    cl = get_column_letter(col_idx)
    if val is not None:
        row24_cells[cl] = str(val).strip()

for col_letter in sorted(row24_cells.keys(), key=lambda x: (len(x), x)):
    print(f"  {col_letter}: {repr(row24_cells[col_letter])}")

# ====== 前5行和最后3行数据 ======
print(f"\n{'=' * 80}")
print(f"【5. 数据行示例】")
print(f"{'=' * 80}")

def print_row(row_num, label=""):
    cells = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_num, column=col_idx).value
        cl = get_column_letter(col_idx)
        if val is not None:
            cells[cl] = str(val).strip()[:80]
    if cells:
        print(f"\n  第{row_num}行 {label}:")
        for cl in sorted(cells.keys(), key=lambda x: (len(x), x)):
            print(f"    {cl}: {repr(cells[cl])}")
    else:
        print(f"\n  第{row_num}行 {label}: (空行)")

# 打印数据行 (从第26行开始)
data_start = 26
data_end = ws.max_row

# 前10行数据
print(f"\n--- 前10行数据 (第{data_start}行~第{data_start+9}行) ---")
for r in range(data_start, min(data_start + 10, data_end + 1)):
    print_row(r)

# 中间3行
mid = (data_start + data_end) // 2
print(f"\n--- 中间3行数据 (第{mid-1}行~第{mid+1}行) ---")
for r in range(max(data_start, mid - 1), min(mid + 2, data_end + 1)):
    print_row(r)

# 最后5行
print(f"\n--- 最后5行数据 (第{data_end-4}行~第{data_end}行) ---")
for r in range(max(data_start, data_end - 4), data_end + 1):
    print_row(r)

# ====== 列分类分析 ======
print(f"\n{'=' * 80}")
print(f"【6. 列分类分析】")
print(f"{'=' * 80}")

# 获取第24行作为最终列头
final_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=24, column=col_idx).value
    if val is not None:
        final_headers[get_column_letter(col_idx)] = str(val).strip()

# 获取第23行作为辅助列头
aux_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=23, column=col_idx).value
    if val is not None:
        aux_headers[get_column_letter(col_idx)] = str(val).strip()

# 获取第7行作为大分组
group7 = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=7, column=col_idx).value
    if val is not None:
        group7[get_column_letter(col_idx)] = str(val).strip()

print(f"\n第7行(大分组):")
for cl in sorted(group7.keys(), key=lambda x: (len(x), x)):
    print(f"  {cl}: {repr(group7[cl])}")

print(f"\n第23行(分组):")
for cl in sorted(aux_headers.keys(), key=lambda x: (len(x), x)):
    print(f"  {cl}: {repr(aux_headers[cl])}")

print(f"\n第24行(列头):")
for cl in sorted(final_headers.keys(), key=lambda x: (len(x), x)):
    print(f"  {cl}: {repr(final_headers[cl])}")

# ====== 合并单元格影响分析 ======
print(f"\n{'=' * 80}")
print(f"【7. 合并单元格影响分析】")
print(f"{'=' * 80}")

# 检查每个合并单元格，看其范围内的数据
for mc in sorted(ws.merged_cells.ranges, key=str):
    min_col = mc.min_col
    max_col = mc.max_col
    min_row = mc.min_row
    max_row = mc.max_row
    top_left = ws.cell(min_row, min_col).value
    # 只有合并区域大于1才展示
    if (max_row - min_row > 0) or (max_col - min_col > 0):
        other_vals = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                v = ws.cell(r, c).value
                if v is not None and (r != min_row or c != min_col):
                    other_vals.append((r, c, str(v)[:40]))
        print(f"  合并区域: {mc}, 左上值: {repr(str(top_left)[:60])}")
        if other_vals:
            for r, c, v in other_vals:
                print(f"    单元格({r},{c})也有值: {repr(v)}")

# ====== 数据完整性检查 ======
print(f"\n{'=' * 80}")
print(f"【8. 数据完整性检查】")
print(f"{'=' * 80}")

# 检查每列空值率
total_data_rows = data_end - data_start + 1
print(f"\n数据行数: {total_data_rows}")
print(f"\n每列非空统计:")
for col_idx in range(1, ws.max_column + 1):
    cl = get_column_letter(col_idx)
    non_null = 0
    sample_vals = set()
    for row_idx in range(data_start, data_end + 1):
        v = ws.cell(row=row_idx, column=col_idx).value
        if v is not None:
            non_null += 1
            s = str(v).strip()[:30]
            if len(s) > 0:
                sample_vals.add(s)
    header = final_headers.get(cl, f"列{cl}")
    rate = non_null / total_data_rows * 100
    samples = list(sample_vals)[:5]
    print(f"  {cl:4s} {header:20s} 非空: {non_null:4d}/{total_data_rows} ({rate:5.1f}%)  样例: {samples}")

# 保存json供后续分析
result = {
    "file": FILE,
    "sheet_name": ws.title,
    "max_row": ws.max_row,
    "max_column": ws.max_column,
    "merged_cells": [str(m) for m in ws.merged_cells.ranges],
    "col_headers_raw": {k: v for k, v in col_headers.items()},
    "row24_headers": row24_cells,
}
with open("F:/yatushi-6-6/scan_uv_oil_result.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\n扫描完成！结果已保存到 scan_uv_oil_result.json")
wb.close()
